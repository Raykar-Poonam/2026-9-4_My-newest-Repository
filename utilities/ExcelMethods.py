import openpyxl

def numRows(file_name,sheet_name):
    Excel_File = openpyxl.load_workbook(file_name)
    Sheet = Excel_File[sheet_name]
    return Sheet.max_row

def readData(file_name,sheet_name,row_num,col_num):
    Excel_File = openpyxl.load_workbook(file_name)
    Sheet = Excel_File[sheet_name]
    return Sheet.cell(row=row_num,column=col_num).value

def writeData(file_name,sheet_name,row_num,col_num,data):
    Excel_File = openpyxl.load_workbook(file_name)
    Sheet = Excel_File[sheet_name]
    Sheet.cell(row=row_num,column=col_num).value=data
    Excel_File.save(file_name)